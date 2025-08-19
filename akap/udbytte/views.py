# SPDX-FileCopyrightText: 2023 Magenta ApS <info@magenta.dk>
#
# SPDX-License-Identifier: MPL-2.0

import csv
import datetime
import json
import logging
import os
from io import StringIO
from typing import List

from aka.utils import AKAJSONEncoder, gettext_lang, send_mail
from django.conf import settings
from django.forms import model_to_dict
from django.template.response import TemplateResponse
from django.views.generic.edit import CreateView
from openpyxl import Workbook, load_workbook
from project.util import split_postnr_by
from project.view_mixins import ErrorHandlerMixin, IsContentMixin, PdfRendererMixin
from udbytte.forms import UdbytteForm, UdbytteFormSet
from udbytte.models import U1A, U1AItem

logger = logging.getLogger(__name__)


class UdbytteCreateView(
    PdfRendererMixin, IsContentMixin, ErrorHandlerMixin, CreateView
):
    model = U1A
    form_class = UdbytteForm
    template_name = "udbytte/form.html"
    pdf_template_name = "udbytte/form.html"
    pdf_css_files = ["css/pdf.css"]

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.object = None
        self.items: U1AItem | None = None

    def get_csv(self):
        # Payment Year;"Payer CVR (udbetaler)";"Recipient CPR (modtager)"; Amount
        csv_io = StringIO()
        writer = csv.writer(csv_io, delimiter=";")
        for item in self.items:
            writer.writerow(
                [
                    self.object.regnskabsår,
                    self.object.cvr,
                    item.cpr_cvr_tin,
                    item.udbytte,
                ]
            )
        return csv_io.getvalue()

    def send_mail_to_submitter(self, pdf_data):
        subject = " / ".join(
            [
                gettext_lang("kl", "udbytte.mail1.subject").format(
                    company_name=self.object.virksomhedsnavn,
                    year=self.object.regnskabsår,
                ),
                gettext_lang("da", "udbytte.mail1.subject").format(
                    company_name=self.object.virksomhedsnavn,
                    year=self.object.regnskabsår,
                ),
            ]
        )
        textbody = [gettext_lang("kl", "udbytte.mail1.textbody")]
        if not self.object.u1_udfyldt:
            textbody.append(
                gettext_lang("kl", "udbytte.mail1.textreminder").format(
                    url=settings.TAX_FORM_U1
                )
            )
        textbody.append(gettext_lang("da", "udbytte.mail1.textbody"))
        if not self.object.u1_udfyldt:
            textbody.append(
                gettext_lang("da", "udbytte.mail1.textreminder").format(
                    url=settings.TAX_FORM_U1
                )
            )

        htmlbody = ["<html><body>", gettext_lang("kl", "udbytte.mail1.htmlbody")]
        if not self.object.u1_udfyldt:
            htmlbody.append(
                gettext_lang("kl", "udbytte.mail1.htmlreminder").format(
                    url=settings.TAX_FORM_U1
                )
            )
        htmlbody.append(gettext_lang("da", "udbytte.mail1.htmlbody"))
        if not self.object.u1_udfyldt:
            htmlbody.append(
                gettext_lang("da", "udbytte.mail1.htmlreminder").format(
                    url=settings.TAX_FORM_U1
                )
            )
        htmlbody.append("</body></html>")

        send_mail(
            recipient=self.object.email,
            subject=subject,
            textbody="\n".join(textbody),
            htmlbody="\n".join(htmlbody),
            attachments=(("formulardata.pdf", pdf_data, "application/pdf"),),
        )

    def send_mail_to_office(self, csv_data, pdf_data):
        subject = " / ".join(
            [
                gettext_lang("kl", "udbytte.mail2.subject").format(
                    company_name=self.object.virksomhedsnavn,
                    year=self.object.regnskabsår,
                ),
                gettext_lang("da", "udbytte.mail2.subject").format(
                    company_name=self.object.virksomhedsnavn,
                    year=self.object.regnskabsår,
                ),
            ]
        )
        textbody = "\n\n".join(
            [
                gettext_lang("kl", "udbytte.mail2.textbody").format(
                    company_name=self.object.virksomhedsnavn,
                    csv=csv_data,
                ),
                gettext_lang("da", "udbytte.mail2.textbody").format(
                    company_name=self.object.virksomhedsnavn,
                    csv=csv_data,
                ),
            ]
        )
        htmlbody = (
            "<html><body><p>" + textbody.replace("\n", "<br/>") + "</body></html>"
        )
        send_mail(
            recipient=settings.EMAIL_OFFICE_RECIPIENT,
            subject=subject,
            textbody=textbody,
            htmlbody=htmlbody,
            attachments=(("formulardata.pdf", pdf_data, "application/pdf"),),
        )

    def render_filled_form(self):
        self.is_pdf = True
        return self.render(
            context=self.get_context_data(
                form=UdbytteForm(instance=self.object),
                formset=UdbytteFormSet(instance=self.object),
                pdf=True,
            ),
            wrap_in_response=False,
        )

    def save_files(self, csv_data, pdf_data):
        folder = f"{settings.TAX_FORM_STORAGE}/{self.object.regnskabsår}/{self.object.dato}/{self.object.cvr}"
        os.makedirs(folder, exist_ok=True)
        file_base_name = f"{datetime.datetime.now().isoformat()}"

        with open(f"{folder}/{file_base_name}.pdf", "wb") as file:
            file.write(pdf_data)

        data = {
            "form": model_to_dict(self.object),
            "formset": [model_to_dict(item) for item in self.items],
            "csv": csv_data,
        }
        with open(f"{folder}/{file_base_name}.json", "w") as file:
            file.write(json.dumps(data, indent=2, cls=AKAJSONEncoder))

    def get_context_data(self, **kwargs):
        return super().get_context_data(
            **{
                **kwargs,
                "u1_url_wrapped": json.dumps({"url": settings.TAX_FORM_U1}),
                "formset": self.get_formset(),
            }
        )

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["initial"]["dato"] = datetime.date.today().strftime("%d/%m/%Y")
        kwargs["oprettet_af_cpr"] = self.request.session["user_info"]["cpr"]
        return kwargs

    def get_formset_kwargs(self):
        return super().get_form_kwargs()

    def get_formset(self):
        return UdbytteFormSet(**self.get_formset_kwargs())

    def post(self, request, *args, **kwargs):
        self.object = None
        form = self.get_form()
        form.full_clean()
        if form.is_valid():
            if form.cleaned_data["use_file"]:
                return self.form_valid(form, None)
            formset = self.get_formset()
            # Trigger form cleaning of both form and formset
            # This ensures that:
            # * we have cleaned_data for the cross-check in clean_with_formset, which validates data across both forms,
            # * all errors are collected for display in form_invalid, not just from the first form that fails
            formset.full_clean()
            # compares data between form and formset, and adds any errors to form
            form.clean_with_formset(formset)
            if formset.is_valid():
                return self.form_valid(form, formset)
        return self.form_invalid(form, formset)

    def form_valid(self, form, formset):
        self.object = form.save(True)
        if form.cleaned_data["use_file"]:
            self.items = self.load_file(form.cleaned_data["file"], self.object)
        elif formset is not None:
            formset.instance = self.object
            self.items = formset.save()

        # PDF handling
        pdf_data = self.render_filled_form()
        self.send_mail_to_submitter(pdf_data)

        # CSV handling
        csv_data = self.get_csv()
        self.save_files(csv_data, pdf_data)
        self.send_mail_to_office(csv_data, pdf_data)

        return TemplateResponse(
            request=self.request,
            template="udbytte/success.html",
            context={},
            using=self.template_engine,
        )

    def form_invalid(self, form, formset):
        return self.render_to_response(
            self.get_context_data(form=form, formset=formset)
        )

    def load_file(self, file, object: U1A) -> List[U1AItem]:
        workbook: Workbook = load_workbook(file)
        items = []
        for sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            required_headers = {
                "generalforsamlingsdato",
                "udbetalingsdato",
                "identifikation",
                "navn",
                "c/o",
                "adresse",
                "postnr.",
                "land",
                "bruttoudbytte",
            }
            headers: List[str] = []
            for row in sheet.iter_rows():
                data = [cell.value for cell in row]
                if len(headers) == 0:
                    pruned_data = [
                        str(x).strip().lower() for x in data if x is not None
                    ]
                    if required_headers.issubset(set(pruned_data)):
                        headers = pruned_data
                else:
                    d = dict(zip(headers, data))
                    if d["identifikation"] in (None, ""):
                        continue
                    postnr_by = split_postnr_by(d["postnr."])
                    if postnr_by is None:
                        postnr = ""
                        by = d["postnr."]
                    else:
                        postnr, by = postnr_by
                    items.append(
                        U1AItem.objects.create(
                            u1a=object,
                            cpr_cvr_tin=d["identifikation"],
                            navn=d["navn"],
                            adresse=d["adresse"],
                            postnummer=postnr,
                            by=by,
                            land=d["land"],
                            udbytte=d["bruttoudbytte"],
                            oprettet=d["udbetalingsdato"],
                        )
                    )
            if len(headers) == 0:
                print(
                    f"Did not find a row that contains all required headers ({required_headers})"
                )
        return items
